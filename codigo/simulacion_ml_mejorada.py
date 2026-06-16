from __future__ import annotations

import json
import math
import re
import unicodedata
import warnings
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
import joblib
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pandas.tseries.offsets import BDay
from sklearn.ensemble import ExtraTreesRegressor, HistGradientBoostingRegressor, RandomForestRegressor
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.model_selection import TimeSeriesSplit
from xgboost import XGBRegressor

warnings.filterwarnings("ignore")
sns.set_theme(style="whitegrid")

try:
    import optuna

    optuna.logging.set_verbosity(optuna.logging.WARNING)
    HAS_OPTUNA = True
except Exception:
    optuna = None
    HAS_OPTUNA = False

try:
    from catboost import CatBoostRegressor

    HAS_CATBOOST = True
except Exception:
    CatBoostRegressor = None
    HAS_CATBOOST = False

try:
    import shap

    HAS_SHAP = True
except Exception:
    shap = None
    HAS_SHAP = False

try:
    from lightgbm import LGBMRegressor

    HAS_LIGHTGBM = True
except Exception:
    LGBMRegressor = None
    HAS_LIGHTGBM = False

try:
    from statsmodels.tsa.statespace.sarimax import SARIMAX

    HAS_STATSMODELS = True
except Exception:
    SARIMAX = None
    HAS_STATSMODELS = False

try:
    from statsmodels.tools.sm_exceptions import ConvergenceWarning

    warnings.filterwarnings("ignore", category=ConvergenceWarning)
except Exception:
    pass


RANDOM_STATE = 42
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"
RESULTS_DIR = PROJECT_DIR / "resultados"
FIGURES_DIR = RESULTS_DIR / "graficas"
MODELS_DIR = RESULTS_DIR / "modelos"
LATEX_DIR = PROJECT_DIR / "latex"
PDF_DIR = PROJECT_DIR / "pdf"
for directory in [DATA_DIR, RESULTS_DIR, FIGURES_DIR, MODELS_DIR, LATEX_DIR, PDF_DIR]:
    directory.mkdir(parents=True, exist_ok=True)

DEFAULT_DATA_FILE = DATA_DIR / "historicoauplata.xlsx"

OUTPUT_EXCEL = RESULTS_DIR / "resultados_modelo_oro_mejorado.xlsx"
PREDICTIONS_CSV = RESULTS_DIR / "predicciones_test_modelo_mejorado.csv"
REPORT_TXT = RESULTS_DIR / "reporte_modelo_oro_mejorado.txt"
MODEL_FILE = MODELS_DIR / "modelo_ganador_oro.joblib"
PLOT_SERIES = FIGURES_DIR / "01_eda_serie_oro.png"
PLOT_PRED = FIGURES_DIR / "02_prediccion_vs_real_test.png"
PLOT_ERRORS = FIGURES_DIR / "03_errores_test.png"
PLOT_IMPORTANCE = FIGURES_DIR / "04_importancia_variables.png"
PLOT_SHAP = FIGURES_DIR / "05_shap_summary.png"
FORECAST_2026_CSV = RESULTS_DIR / "pronostico_mayo_junio_julio_2026.csv"
FORECAST_2035_CSV = RESULTS_DIR / "pronostico_hasta_2035.csv"
SINGLE_TRAJECTORY_CSV = RESULTS_DIR / "trayectoria_unica_monte_carlo_base.csv"
PLOT_FORECAST_2026 = FIGURES_DIR / "06_pronostico_mayo_junio_julio_2026.png"
PLOT_FORECAST_2035 = FIGURES_DIR / "07_pronostico_hasta_2035.png"
PLOT_FORECAST_2035_PATHS = FIGURES_DIR / "08_trayectorias_monte_carlo_2035.png"
PLOT_SINGLE_TRAJECTORY = FIGURES_DIR / "09_trayectoria_unica_monte_carlo_base.png"

SHORT_FORECAST_START = pd.Timestamp("2026-05-01")
SHORT_FORECAST_END = pd.Timestamp("2026-07-31")
LONG_FORECAST_END = pd.Timestamp("2035-12-31")
MONTE_CARLO_SIMULATIONS = 120
MONTE_CARLO_SAMPLE_PATHS = 12
OPTUNA_TRIALS = 4
ENABLE_SARIMAX_REFERENCE = False

FRED_SERIES = {
    "DGS10": "treasury_10y",
    "T10YIE": "breakeven_10y",
    "CPIAUCSL": "cpi_us",
    "DTWEXBGS": "dollar_index",
    "FEDFUNDS": "fed_funds",
    "VIXCLS": "vix",
    "DCOILWTICO": "wti_oil",
    "SP500": "sp500",
    "DEXCOUS": "usd_cop",
}

SCENARIO_ORDER = ["base", "crisis_alta_incertidumbre", "optimista"]
SCENARIO_COLORS = {
    "base": "darkorange",
    "crisis_alta_incertidumbre": "darkgreen",
    "optimista": "steelblue",
}


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def first_existing_file() -> Path:
    candidates = [
        DEFAULT_DATA_FILE,
        DATA_DIR / "historicoauplata.xls",
        DATA_DIR / "historicoauplata.csv",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError("No se encontro historicoauplata en la carpeta data del proyecto.")


def detect_header_row_excel(file_path: Path, sheet_name: str, max_rows: int = 10) -> int:
    preview = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=max_rows)
    best_row = 0
    best_score = -1

    for row_idx in range(len(preview)):
        values = [normalize_text(value) for value in preview.iloc[row_idx].tolist() if pd.notna(value)]
        if not values:
            continue

        score = 0
        if any(token in {"fecha", "date", "fecha_de_publicacion"} for token in values):
            score += 5
        if any(token in {"cierre", "close"} for token in values):
            score += 5
        if any("oro" in token or "gold" in token or "xau" in token for token in values):
            score += 3
        if any("precio" in token or "price" in token for token in values):
            score += 2
        if any(token in {"apertura", "open", "maximo", "high", "minimo", "low"} for token in values):
            score += 2

        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def read_excel_sheet_auto(file_path: Path, sheet_name: str) -> pd.DataFrame:
    header_row = detect_header_row_excel(file_path, sheet_name)
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
    df = df.dropna(how="all").copy()
    df.columns = [str(col).strip() for col in df.columns]
    return df


def detect_date_column(columns: list[str]) -> str:
    ranked = []
    for col in columns:
        norm = normalize_text(col)
        score = 0
        if norm == "fecha":
            score += 10
        if "fecha" in norm or "date" in norm:
            score += 6
        if "publicacion" in norm:
            score += 1
        ranked.append((score, col))

    ranked.sort(reverse=True, key=lambda item: item[0])
    if not ranked or ranked[0][0] <= 0:
        raise ValueError("No se pudo detectar una columna de fecha.")
    return ranked[0][1]


def detect_gold_target_column(columns: list[str]) -> str:
    best_col = None
    best_score = -10

    for col in columns:
        norm = normalize_text(col)
        score = 0
        if norm in {"cierre", "close"}:
            score += 12
        if "cierre" in norm or "close" in norm:
            score += 8
        if "oro" in norm or "gold" in norm or "xau" in norm:
            score += 5
        if "precio" in norm or "price" in norm:
            score += 4
        if any(word in norm for word in ["apertura", "open", "maximo", "high", "minimo", "low"]):
            score -= 6
        if any(word in norm for word in ["plata", "silver", "recurso", "reserva", "resource", "reserve"]):
            score -= 6
        if any(word in norm for word in ["fecha", "date", "fuente", "url"]):
            score -= 8

        if score > best_score:
            best_score = score
            best_col = col

    if best_col is None or best_score <= 0:
        raise ValueError("No se pudo detectar la variable objetivo del oro.")
    return best_col


def detect_ohlc_columns(columns: list[str]) -> dict[str, str]:
    mapping = {}
    keywords = {
        "apertura": ["apertura", "open"],
        "maximo": ["maximo", "high"],
        "minimo": ["minimo", "low"],
        "cierre": ["cierre", "close"],
    }

    for canonical_name, options in keywords.items():
        for col in columns:
            norm = normalize_text(col)
            if any(option == norm or option in norm for option in options):
                mapping[canonical_name] = col
                break

    return mapping


def score_gold_sheet(file_path: Path, sheet_name: str) -> tuple[float, pd.DataFrame]:
    df = read_excel_sheet_auto(file_path, sheet_name)
    if df.empty:
        return -1e9, df

    columns = list(df.columns)
    score = 0.0

    try:
        _ = detect_date_column(columns)
        score += 8
    except ValueError:
        return -1e9, df

    try:
        target_col = detect_gold_target_column(columns)
        score += 10
        target_norm = normalize_text(target_col)
        if "cierre" in target_norm or "close" in target_norm:
            score += 5
    except ValueError:
        return -1e9, df

    score += min(len(df) / 100.0, 30.0)
    name_norm = normalize_text(sheet_name)
    if "histor" in name_norm:
        score += 10
    if name_norm.isdigit():
        score += 2
    score += len(detect_ohlc_columns(columns))

    return score, df


def load_gold_data(file_path: Path) -> tuple[pd.DataFrame, dict[str, str]]:
    if file_path.suffix.lower() == ".csv":
        df = pd.read_csv(file_path)
        selected_sheet = file_path.name
    else:
        workbook = pd.ExcelFile(file_path)
        best_sheet = None
        best_df = None
        best_score = -1e9
        for sheet_name in workbook.sheet_names:
            score, df = score_gold_sheet(file_path, sheet_name)
            if score > best_score:
                best_score = score
                best_sheet = sheet_name
                best_df = df
        if best_df is None or best_sheet is None:
            raise ValueError("No se encontro una hoja adecuada con historicos del oro.")
        df = best_df
        selected_sheet = best_sheet

    date_col = detect_date_column(list(df.columns))
    target_col = detect_gold_target_column(list(df.columns))
    ohlc_cols = detect_ohlc_columns(list(df.columns))

    rename_map = {date_col: "fecha", target_col: "precio_oro"}
    for canonical_name, original_name in ohlc_cols.items():
        if original_name != target_col:
            rename_map[original_name] = canonical_name

    df = df.rename(columns=rename_map).copy()
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    df["precio_oro"] = pd.to_numeric(df["precio_oro"], errors="coerce")
    for col in ["apertura", "maximo", "minimo"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    keep_cols = ["fecha", "precio_oro", "apertura", "maximo", "minimo"]
    existing_cols = [col for col in keep_cols if col in df.columns]
    df = (
        df[existing_cols]
        .dropna(subset=["fecha", "precio_oro"])
        .drop_duplicates(subset=["fecha"], keep="last")
        .sort_values("fecha")
        .reset_index(drop=True)
    )
    df = df[df["fecha"] >= pd.Timestamp("2016-01-01")].reset_index(drop=True)

    metadata = {
        "archivo": str(file_path),
        "hoja_oro": selected_sheet,
        "columna_fecha": date_col,
        "columna_objetivo": target_col,
    }
    return df, metadata


def find_sheet_by_keyword(file_path: Path, keywords: list[str]) -> str | None:
    if file_path.suffix.lower() == ".csv":
        return None

    workbook = pd.ExcelFile(file_path)
    normalized_keywords = [normalize_text(keyword) for keyword in keywords]
    for sheet_name in workbook.sheet_names:
        name_norm = normalize_text(sheet_name)
        if any(keyword in name_norm for keyword in normalized_keywords):
            return sheet_name
    return None


def load_silver_data(file_path: Path) -> tuple[pd.DataFrame | None, dict[str, str]]:
    sheet_name = find_sheet_by_keyword(file_path, ["plata", "silver", "ag"])
    if sheet_name is None:
        return None, {}

    df = read_excel_sheet_auto(file_path, sheet_name)
    date_col = detect_date_column(list(df.columns))

    price_col = None
    best_score = -1
    for col in df.columns:
        norm = normalize_text(col)
        score = 0
        if "plata" in norm or "silver" in norm or norm == "ag":
            score += 8
        if "precio" in norm or "price" in norm:
            score += 4
        if any(word in norm for word in ["fecha", "date", "fuente", "url"]):
            score -= 8
        if score > best_score:
            best_score = score
            price_col = col

    if price_col is None or best_score <= 0:
        return None, {}

    silver = df.rename(columns={date_col: "fecha", price_col: "precio_plata"}).copy()
    silver["fecha"] = pd.to_datetime(silver["fecha"], errors="coerce")
    silver["precio_plata"] = pd.to_numeric(silver["precio_plata"], errors="coerce")
    silver = (
        silver[["fecha", "precio_plata"]]
        .dropna(subset=["fecha", "precio_plata"])
        .drop_duplicates(subset=["fecha"], keep="last")
        .sort_values("fecha")
        .reset_index(drop=True)
    )

    metadata = {
        "hoja_plata": sheet_name,
        "columna_plata": price_col,
    }
    return silver, metadata


def load_ni_data(file_path: Path) -> tuple[pd.DataFrame | None, dict[str, str]]:
    sheet_name = find_sheet_by_keyword(file_path, ["ni"])
    if sheet_name is None:
        return None, {}

    df = read_excel_sheet_auto(file_path, sheet_name)
    columns = list(df.columns)

    try:
        date_col = detect_date_column(columns)
    except ValueError:
        return None, {}

    resource_col = None
    reserve_col = None
    for col in columns:
        norm = normalize_text(col)
        if "recurso" in norm or "resource" in norm:
            resource_col = col
        if "reserva" in norm or "reserve" in norm:
            reserve_col = col

    if resource_col is None and reserve_col is None:
        return None, {}

    rename_map = {date_col: "fecha_publicacion"}
    if resource_col is not None:
        rename_map[resource_col] = "ni_precio_recursos"
    if reserve_col is not None:
        rename_map[reserve_col] = "ni_precio_reservas"

    ni = df.rename(columns=rename_map).copy()
    ni["fecha_publicacion"] = pd.to_datetime(ni["fecha_publicacion"], errors="coerce")
    for col in ["ni_precio_recursos", "ni_precio_reservas"]:
        if col in ni.columns:
            ni[col] = pd.to_numeric(ni[col], errors="coerce")

    keep_cols = ["fecha_publicacion", "ni_precio_recursos", "ni_precio_reservas"]
    keep_cols = [col for col in keep_cols if col in ni.columns]
    ni = (
        ni[keep_cols]
        .dropna(subset=["fecha_publicacion"])
        .drop_duplicates(subset=["fecha_publicacion"], keep="last")
        .sort_values("fecha_publicacion")
        .reset_index(drop=True)
    )

    metadata = {
        "hoja_ni": sheet_name,
        "columna_fecha_publicacion": date_col,
    }
    return ni, metadata


def fetch_fred_macro_data() -> tuple[pd.DataFrame, dict[str, str], list[str]]:
    macro_frames = []
    metadata = {}
    warnings_list = []

    for series_id, col_name in FRED_SERIES.items():
        try:
            url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}"
            series_df = pd.read_csv(url)
            series_df = series_df.rename(columns={"observation_date": "fecha", series_id: col_name})
            series_df["fecha"] = pd.to_datetime(series_df["fecha"], errors="coerce")
            series_df[col_name] = pd.to_numeric(series_df[col_name], errors="coerce")
            series_df = (
                series_df[["fecha", col_name]]
                .dropna(subset=["fecha"])
                .sort_values("fecha")
                .reset_index(drop=True)
            )
            macro_frames.append(series_df)
            metadata[f"fred_{col_name}"] = series_id
        except Exception as exc:
            warnings_list.append(f"No se pudo descargar {series_id}: {exc}")

    if not macro_frames:
        return pd.DataFrame(columns=["fecha"]), metadata, warnings_list

    macro_df = macro_frames[0]
    for next_df in macro_frames[1:]:
        macro_df = macro_df.merge(next_df, on="fecha", how="outer")

    macro_df = macro_df.sort_values("fecha").reset_index(drop=True)
    if {"treasury_10y", "breakeven_10y"}.issubset(macro_df.columns):
        macro_df["tasa_real_aprox"] = macro_df["treasury_10y"] - macro_df["breakeven_10y"]
    return macro_df, metadata, warnings_list


def build_eda_table(df: pd.DataFrame) -> pd.DataFrame:
    numeric_part = df.select_dtypes(include=[np.number]).describe().T
    datetime_cols = df.select_dtypes(include=["datetime"]).columns.tolist()

    datetime_rows = []
    for col in datetime_cols:
        series = df[col].dropna()
        datetime_rows.append(
            {
                "variable": col,
                "count": float(series.shape[0]),
                "min": series.min(),
                "max": series.max(),
            }
        )

    datetime_part = pd.DataFrame(datetime_rows).set_index("variable") if datetime_rows else pd.DataFrame()
    if not datetime_part.empty:
        eda_table = pd.concat([numeric_part, datetime_part], axis=0, sort=False)
    else:
        eda_table = numeric_part

    eda_table.index.name = "variable"
    return eda_table


def print_basic_eda(df_gold: pd.DataFrame, metadata: dict[str, str]) -> None:
    print("\n=== ANALISIS EXPLORATORIO BASICO ===")
    print(f"Archivo detectado: {metadata['archivo']}")
    print(f"Hoja detectada para oro: {metadata['hoja_oro']}")
    print(f"Columna objetivo detectada: {metadata['columna_objetivo']}")
    print(f"Observaciones: {len(df_gold)}")
    print(f"Rango temporal: {df_gold['fecha'].min():%Y-%m-%d} a {df_gold['fecha'].max():%Y-%m-%d}")
    print("Valores faltantes por columna:")
    print(df_gold.isna().sum().to_string())


def create_basic_eda_plot(df_gold: pd.DataFrame) -> Path:
    fig, ax = plt.subplots(figsize=(13, 5))
    ax.plot(df_gold["fecha"], df_gold["precio_oro"], color="goldenrod", linewidth=1.3)
    ax.set_title("Serie historica del precio del oro utilizada en el modelo")
    ax.set_xlabel("Fecha")
    ax.set_ylabel("Precio del oro (USD/oz)")
    ax.grid(alpha=0.25)
    fig.tight_layout()
    fig.savefig(PLOT_SERIES, dpi=300)
    plt.close(fig)
    return PLOT_SERIES


def assemble_base_timeseries(
    df_gold: pd.DataFrame,
    df_silver: pd.DataFrame | None = None,
    df_ni: pd.DataFrame | None = None,
    df_macro: pd.DataFrame | None = None,
) -> tuple[pd.DataFrame, list[str]]:
    df = df_gold[["fecha", "precio_oro"]].copy().sort_values("fecha").reset_index(drop=True)
    used_exogenous = []

    if df_silver is not None and not df_silver.empty:
        df = df.merge(df_silver, on="fecha", how="left")
        df["precio_plata"] = pd.to_numeric(df["precio_plata"], errors="coerce").ffill()
        if df["precio_plata"].notna().mean() >= 0.50:
            used_exogenous.append("precio_plata")
        else:
            df = df.drop(columns=["precio_plata"])

    if df_ni is not None and not df_ni.empty:
        df = pd.merge_asof(
            df.sort_values("fecha"),
            df_ni.sort_values("fecha_publicacion"),
            left_on="fecha",
            right_on="fecha_publicacion",
            direction="backward",
        )
        if "fecha_publicacion" in df.columns:
            df = df.drop(columns=["fecha_publicacion"])
        for col in ["ni_precio_recursos", "ni_precio_reservas"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").ffill()
                if df[col].notna().mean() >= 0.50:
                    used_exogenous.append(col)
                else:
                    df = df.drop(columns=[col])

    if df_macro is not None and not df_macro.empty:
        df = df.merge(df_macro, on="fecha", how="left")
        macro_cols = [col for col in df_macro.columns if col != "fecha"]
        for col in macro_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").ffill()
                if df[col].notna().mean() >= 0.40:
                    used_exogenous.append(col)
                else:
                    df = df.drop(columns=[col])

    df = df.drop_duplicates(subset=["fecha"], keep="last").sort_values("fecha").reset_index(drop=True)
    return df, sorted(set(used_exogenous))


def engineer_features(df: pd.DataFrame, add_targets: bool = True) -> pd.DataFrame:
    df = df.copy().sort_values("fecha").reset_index(drop=True)

    df["retorno_oro_1"] = df["precio_oro"].pct_change(1)
    df["retorno_oro_5"] = df["precio_oro"].pct_change(5)
    df["retorno_oro_10"] = df["precio_oro"].pct_change(10)
    df["diferencia_oro_1"] = df["precio_oro"].diff(1)
    df["diferencia_oro_5"] = df["precio_oro"].diff(5)
    df["diferencia_oro_10"] = df["precio_oro"].diff(10)

    for lag in [1, 2, 3, 5, 10, 20]:
        df[f"precio_oro_lag_{lag}"] = df["precio_oro"].shift(lag)

    for window in [3, 5, 10, 20, 30]:
        df[f"ma_oro_{window}"] = df["precio_oro"].rolling(window).mean()
        df[f"ema_oro_{window}"] = df["precio_oro"].ewm(span=window, adjust=False).mean()
        df[f"volatilidad_oro_{window}"] = df["retorno_oro_1"].rolling(window).std()

    if "precio_plata" in df.columns:
        df["ratio_oro_plata"] = df["precio_oro"] / df["precio_plata"]
        df["retorno_plata_1"] = df["precio_plata"].pct_change(1)
        df["retorno_plata_5"] = df["precio_plata"].pct_change(5)
        df["diferencia_plata_1"] = df["precio_plata"].diff(1)
        for lag in [1, 2, 5, 10]:
            df[f"precio_plata_lag_{lag}"] = df["precio_plata"].shift(lag)
        for window in [5, 10, 20]:
            df[f"ma_plata_{window}"] = df["precio_plata"].rolling(window).mean()
            df[f"volatilidad_plata_{window}"] = df["retorno_plata_1"].rolling(window).std()

    macro_feature_cols = [
        "treasury_10y",
        "breakeven_10y",
        "cpi_us",
        "tasa_real_aprox",
        "dollar_index",
        "fed_funds",
        "vix",
        "wti_oil",
        "sp500",
        "usd_cop",
    ]
    for col in macro_feature_cols:
        if col in df.columns:
            df[f"{col}_chg_1"] = df[col].pct_change(1)
            df[f"{col}_chg_5"] = df[col].pct_change(5)
            df[f"{col}_lag_1"] = df[col].shift(1)
            df[f"{col}_ma_5"] = df[col].rolling(5).mean()
            df[f"{col}_ma_20"] = df[col].rolling(20).mean()

    if {"ni_precio_recursos", "ni_precio_reservas"}.issubset(df.columns):
        df["ni_gap_recursos_reservas"] = df["ni_precio_recursos"] - df["ni_precio_reservas"]

    df["anio"] = df["fecha"].dt.year
    df["mes"] = df["fecha"].dt.month
    df["trimestre"] = df["fecha"].dt.quarter
    df["dia_mes"] = df["fecha"].dt.day
    df["dia_semana"] = df["fecha"].dt.dayofweek
    df["semana_anio"] = df["fecha"].dt.isocalendar().week.astype(int)
    df["mes_inicio"] = df["fecha"].dt.is_month_start.astype(int)
    df["mes_fin"] = df["fecha"].dt.is_month_end.astype(int)
    df["mes_sin"] = np.sin(2 * np.pi * df["mes"] / 12.0)
    df["mes_cos"] = np.cos(2 * np.pi * df["mes"] / 12.0)
    df["dia_semana_sin"] = np.sin(2 * np.pi * df["dia_semana"] / 7.0)
    df["dia_semana_cos"] = np.cos(2 * np.pi * df["dia_semana"] / 7.0)

    if add_targets:
        df["target_precio_oro_t1"] = df["precio_oro"].shift(-1)
        df["target_logret_t1"] = np.log(df["target_precio_oro_t1"] / df["precio_oro"])

    return df


def build_model_dataset(
    df_gold: pd.DataFrame,
    df_silver: pd.DataFrame | None = None,
    df_ni: pd.DataFrame | None = None,
    df_macro: pd.DataFrame | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    base_df, used_exogenous = assemble_base_timeseries(df_gold, df_silver, df_ni, df_macro)
    df = engineer_features(base_df, add_targets=True)
    df = df.replace([np.inf, -np.inf], np.nan)

    non_numeric_cols = [
        col
        for col in df.columns
        if col not in {"fecha"}
        and col not in {"target_precio_oro_t1", "target_logret_t1"}
        and not pd.api.types.is_numeric_dtype(df[col])
    ]
    if non_numeric_cols:
        df = df.drop(columns=non_numeric_cols)

    candidate_feature_cols = [
        col for col in df.columns if col not in {"fecha", "target_precio_oro_t1", "target_logret_t1"}
    ]
    high_missing_cols = [col for col in candidate_feature_cols if df[col].isna().mean() > 0.35 and col != "precio_oro"]
    if high_missing_cols:
        df = df.drop(columns=high_missing_cols)

    candidate_feature_cols = [
        col for col in df.columns if col not in {"fecha", "target_precio_oro_t1", "target_logret_t1"}
    ]
    df_model = df.dropna(subset=candidate_feature_cols + ["target_precio_oro_t1", "target_logret_t1"]).copy()
    df_model = df_model.sort_values("fecha").reset_index(drop=True)
    return df_model, base_df, used_exogenous


def temporal_train_test_split(df_model: pd.DataFrame, test_ratio: float = 0.20) -> tuple[pd.DataFrame, pd.DataFrame]:
    test_size = max(120, int(len(df_model) * test_ratio))
    if test_size >= len(df_model):
        raise ValueError("No hay suficientes datos para una particion temporal valida.")
    train_df = df_model.iloc[:-test_size].copy()
    test_df = df_model.iloc[-test_size:].copy()
    return train_df, test_df


def regression_metrics(y_true: pd.Series | np.ndarray, y_pred: np.ndarray) -> dict[str, float]:
    y_true_arr = np.asarray(y_true, dtype=float)
    y_pred_arr = np.asarray(y_pred, dtype=float)
    mask = np.isfinite(y_true_arr) & np.isfinite(y_pred_arr)
    y_true_arr = y_true_arr[mask]
    y_pred_arr = y_pred_arr[mask]
    if len(y_true_arr) == 0:
        return {
            "MAE": float("inf"),
            "RMSE": float("inf"),
            "MAPE_pct": float("inf"),
            "R2": float("-inf"),
        }
    epsilon = 1e-8
    mae = mean_absolute_error(y_true_arr, y_pred_arr)
    rmse = math.sqrt(mean_squared_error(y_true_arr, y_pred_arr))
    mape = float(np.mean(np.abs((y_true_arr - y_pred_arr) / np.maximum(np.abs(y_true_arr), epsilon))) * 100.0)
    r2 = r2_score(y_true_arr, y_pred_arr)
    return {
        "MAE": float(mae),
        "RMSE": float(rmse),
        "MAPE_pct": float(mape),
        "R2": float(r2),
    }


def ml_model_candidates() -> list[dict[str, object]]:
    candidates = [
        {
            "family": "XGBoost",
            "params": {
                "n_estimators": 350,
                "max_depth": 3,
                "learning_rate": 0.04,
                "subsample": 0.90,
                "colsample_bytree": 0.90,
                "min_child_weight": 3,
                "reg_alpha": 0.00,
                "reg_lambda": 1.20,
            },
        },
        {
            "family": "XGBoost",
            "params": {
                "n_estimators": 500,
                "max_depth": 4,
                "learning_rate": 0.03,
                "subsample": 0.85,
                "colsample_bytree": 0.85,
                "min_child_weight": 2,
                "reg_alpha": 0.02,
                "reg_lambda": 1.40,
            },
        },
        {
            "family": "RandomForest",
            "params": {
                "n_estimators": 400,
                "max_depth": 10,
                "min_samples_leaf": 2,
                "max_features": 0.8,
            },
        },
        {
            "family": "RandomForest",
            "params": {
                "n_estimators": 600,
                "max_depth": 14,
                "min_samples_leaf": 1,
                "max_features": 0.7,
            },
        },
        {
            "family": "ExtraTrees",
            "params": {
                "n_estimators": 400,
                "max_depth": 12,
                "min_samples_leaf": 2,
                "max_features": 0.85,
            },
        },
        {
            "family": "ExtraTrees",
            "params": {
                "n_estimators": 600,
                "max_depth": 16,
                "min_samples_leaf": 1,
                "max_features": 0.70,
            },
        },
    ]

    if HAS_LIGHTGBM:
        candidates.extend(
            [
                {
                    "family": "LightGBM",
                    "params": {
                        "n_estimators": 350,
                        "learning_rate": 0.04,
                        "num_leaves": 31,
                        "subsample": 0.90,
                        "colsample_bytree": 0.90,
                        "min_child_samples": 20,
                    },
                },
                {
                    "family": "LightGBM",
                    "params": {
                        "n_estimators": 500,
                        "learning_rate": 0.03,
                        "num_leaves": 24,
                        "subsample": 0.85,
                        "colsample_bytree": 0.85,
                        "min_child_samples": 25,
                    },
                },
            ]
        )

    if HAS_CATBOOST:
        candidates.extend(
            [
                {
                    "family": "CatBoost",
                    "params": {
                        "iterations": 350,
                        "depth": 4,
                        "learning_rate": 0.035,
                        "l2_leaf_reg": 5.0,
                    },
                },
                {
                    "family": "CatBoost",
                    "params": {
                        "iterations": 650,
                        "depth": 5,
                        "learning_rate": 0.025,
                        "l2_leaf_reg": 8.0,
                    },
                },
            ]
        )

    candidates.extend(
        [
            {
                "family": "HistGradientBoosting",
                "params": {
                    "max_iter": 350,
                    "learning_rate": 0.035,
                    "max_leaf_nodes": 31,
                    "l2_regularization": 0.02,
                },
            },
            {
                "family": "HistGradientBoosting",
                "params": {
                    "max_iter": 500,
                    "learning_rate": 0.025,
                    "max_leaf_nodes": 24,
                    "l2_regularization": 0.05,
                },
            },
        ]
    )

    compact_candidates = []
    seen_families = set()
    for candidate in candidates:
        family = candidate["family"]
        if family not in seen_families:
            compact_candidates.append(candidate)
            seen_families.add(family)
    return compact_candidates


def sarimax_candidates() -> list[dict[str, object]]:
    if not HAS_STATSMODELS:
        return []
    return [
        {"family": "SARIMAX", "params": {"order": (1, 0, 0), "trend": "n"}},
        {"family": "SARIMAX", "params": {"order": (1, 0, 1), "trend": "n"}},
        {"family": "SARIMAX", "params": {"order": (2, 0, 1), "trend": "n"}},
    ]


def make_ml_model(family: str, params: dict[str, object]):
    if family == "XGBoost":
        return XGBRegressor(
            objective="reg:squarederror",
            eval_metric="rmse",
            tree_method="hist",
            random_state=RANDOM_STATE,
            n_jobs=4,
            **params,
        )
    if family == "RandomForest":
        return RandomForestRegressor(
            random_state=RANDOM_STATE,
            n_jobs=4,
            **params,
        )
    if family == "ExtraTrees":
        return ExtraTreesRegressor(
            random_state=RANDOM_STATE,
            n_jobs=4,
            **params,
        )
    if family == "LightGBM" and HAS_LIGHTGBM:
        return LGBMRegressor(
            objective="regression",
            random_state=RANDOM_STATE,
            n_jobs=4,
            verbosity=-1,
            **params,
        )
    if family == "CatBoost" and HAS_CATBOOST:
        return CatBoostRegressor(
            loss_function="RMSE",
            random_seed=RANDOM_STATE,
            thread_count=4,
            verbose=False,
            allow_writing_files=False,
            **params,
        )
    if family == "HistGradientBoosting":
        return HistGradientBoostingRegressor(
            random_state=RANDOM_STATE,
            **params,
        )
    raise ValueError(f"Modelo no soportado: {family}")


def select_ml_feature_columns(df_model: pd.DataFrame) -> list[str]:
    return [col for col in df_model.columns if col not in {"fecha", "target_precio_oro_t1", "target_logret_t1"}]


def select_sarimax_exog_columns(df_model: pd.DataFrame) -> list[str]:
    preferred_cols = [
        "precio_plata",
        "precio_plata_lag_1",
        "precio_plata_lag_2",
        "retorno_plata_1",
        "retorno_plata_5",
        "ni_precio_recursos",
        "ni_precio_reservas",
        "ni_gap_recursos_reservas",
        "treasury_10y",
        "treasury_10y_lag_1",
        "treasury_10y_chg_1",
        "breakeven_10y",
        "breakeven_10y_lag_1",
        "breakeven_10y_chg_1",
        "cpi_us",
        "cpi_us_lag_1",
        "cpi_us_chg_1",
        "tasa_real_aprox",
        "tasa_real_aprox_lag_1",
        "dollar_index",
        "dollar_index_lag_1",
        "dollar_index_chg_1",
        "fed_funds",
        "fed_funds_lag_1",
        "vix",
        "vix_lag_1",
        "vix_chg_1",
        "wti_oil",
        "wti_oil_lag_1",
        "wti_oil_chg_1",
        "sp500",
        "sp500_lag_1",
        "sp500_chg_1",
        "usd_cop",
        "usd_cop_lag_1",
        "usd_cop_chg_1",
        "anio",
        "mes",
        "trimestre",
        "dia_semana",
        "mes_inicio",
        "mes_fin",
        "mes_sin",
        "mes_cos",
        "dia_semana_sin",
        "dia_semana_cos",
    ]
    candidate_cols = [col for col in preferred_cols if col in df_model.columns]
    return [col for col in candidate_cols if df_model[col].nunique(dropna=False) > 1]


def optuna_tuned_candidate(train_df: pd.DataFrame, feature_cols: list[str]) -> list[dict[str, object]]:
    if not HAS_OPTUNA:
        return []

    X_train = train_df[feature_cols]
    y_train = train_df["target_logret_t1"]
    families = ["XGBoost", "RandomForest", "ExtraTrees", "HistGradientBoosting"]
    if HAS_LIGHTGBM:
        families.append("LightGBM")
    if HAS_CATBOOST:
        families.append("CatBoost")
    tscv = TimeSeriesSplit(n_splits=3)

    def suggest_params(trial, family: str) -> dict[str, object]:
        if family == "XGBoost":
            return {
                "n_estimators": trial.suggest_int("xgb_n_estimators", 250, 700, step=50),
                "max_depth": trial.suggest_int("xgb_max_depth", 2, 5),
                "learning_rate": trial.suggest_float("xgb_learning_rate", 0.015, 0.07, log=True),
                "subsample": trial.suggest_float("xgb_subsample", 0.70, 0.95),
                "colsample_bytree": trial.suggest_float("xgb_colsample_bytree", 0.70, 0.95),
                "min_child_weight": trial.suggest_int("xgb_min_child_weight", 1, 6),
                "reg_alpha": trial.suggest_float("xgb_reg_alpha", 0.0, 0.08),
                "reg_lambda": trial.suggest_float("xgb_reg_lambda", 0.8, 2.2),
            }
        if family == "LightGBM":
            return {
                "n_estimators": trial.suggest_int("lgb_n_estimators", 250, 700, step=50),
                "learning_rate": trial.suggest_float("lgb_learning_rate", 0.015, 0.07, log=True),
                "num_leaves": trial.suggest_int("lgb_num_leaves", 16, 48),
                "subsample": trial.suggest_float("lgb_subsample", 0.70, 0.95),
                "colsample_bytree": trial.suggest_float("lgb_colsample_bytree", 0.70, 0.95),
                "min_child_samples": trial.suggest_int("lgb_min_child_samples", 12, 35),
            }
        if family == "CatBoost":
            return {
                "iterations": trial.suggest_int("cat_iterations", 300, 750, step=50),
                "depth": trial.suggest_int("cat_depth", 3, 6),
                "learning_rate": trial.suggest_float("cat_learning_rate", 0.015, 0.07, log=True),
                "l2_leaf_reg": trial.suggest_float("cat_l2_leaf_reg", 2.0, 12.0),
            }
        if family == "RandomForest":
            return {
                "n_estimators": trial.suggest_int("rf_n_estimators", 300, 700, step=100),
                "max_depth": trial.suggest_int("rf_max_depth", 8, 18),
                "min_samples_leaf": trial.suggest_int("rf_min_samples_leaf", 1, 4),
                "max_features": trial.suggest_float("rf_max_features", 0.55, 0.90),
            }
        if family == "ExtraTrees":
            return {
                "n_estimators": trial.suggest_int("et_n_estimators", 300, 700, step=100),
                "max_depth": trial.suggest_int("et_max_depth", 8, 20),
                "min_samples_leaf": trial.suggest_int("et_min_samples_leaf", 1, 4),
                "max_features": trial.suggest_float("et_max_features", 0.55, 0.90),
            }
        return {
            "max_iter": trial.suggest_int("hgb_max_iter", 250, 700, step=50),
            "learning_rate": trial.suggest_float("hgb_learning_rate", 0.015, 0.07, log=True),
            "max_leaf_nodes": trial.suggest_int("hgb_max_leaf_nodes", 16, 48),
            "l2_regularization": trial.suggest_float("hgb_l2_regularization", 0.0, 0.12),
        }

    def objective(trial) -> float:
        family = trial.suggest_categorical("family", families)
        params = suggest_params(trial, family)
        rmse_values = []
        for idx_fit, idx_val in tscv.split(X_train):
            model = make_ml_model(family, params)
            model.fit(X_train.iloc[idx_fit], y_train.iloc[idx_fit])
            pred_val_logret = model.predict(X_train.iloc[idx_val])
            base_price_val = train_df.iloc[idx_val]["precio_oro"].to_numpy(dtype=float)
            pred_val_price = base_price_val * np.exp(pred_val_logret)
            y_val_price = train_df.iloc[idx_val]["target_precio_oro_t1"].to_numpy(dtype=float)
            rmse_values.append(regression_metrics(y_val_price, pred_val_price)["RMSE"])
        return float(np.mean(rmse_values))

    try:
        study = optuna.create_study(direction="minimize", sampler=optuna.samplers.TPESampler(seed=RANDOM_STATE))
        study.optimize(objective, n_trials=OPTUNA_TRIALS, show_progress_bar=False)
        best_family = study.best_trial.params["family"]
        best_params = suggest_params(study.best_trial, best_family)
        return [{"family": best_family, "params": best_params, "optimized_by": "Optuna", "optuna_best_rmse": study.best_value}]
    except Exception as exc:
        print(f"Optuna no pudo completar la busqueda automatica: {exc}")
        return []


def cross_validate_ml_models(
    train_df: pd.DataFrame,
    feature_cols: list[str],
) -> tuple[pd.DataFrame, dict[str, object]]:
    X_train = train_df[feature_cols]
    y_train = train_df["target_logret_t1"]
    splits = 3
    tscv = TimeSeriesSplit(n_splits=splits)
    rows = []

    candidates = ml_model_candidates() + optuna_tuned_candidate(train_df, feature_cols)
    for candidate_id, candidate in enumerate(candidates, start=1):
        family = str(candidate["family"])
        params = dict(candidate["params"])
        fold_metrics = []

        for fold_id, (idx_fit, idx_val) in enumerate(tscv.split(X_train), start=1):
            model = make_ml_model(family, params)
            model.fit(X_train.iloc[idx_fit], y_train.iloc[idx_fit])
            pred_val_logret = model.predict(X_train.iloc[idx_val])
            base_price_val = train_df.iloc[idx_val]["precio_oro"].to_numpy(dtype=float)
            pred_val_price = base_price_val * np.exp(pred_val_logret)
            y_val_price = train_df.iloc[idx_val]["target_precio_oro_t1"].to_numpy(dtype=float)
            metrics = regression_metrics(y_val_price, pred_val_price)
            metrics["fold"] = fold_id
            fold_metrics.append(metrics)

        fold_df = pd.DataFrame(fold_metrics)
        rows.append(
            {
                "candidate_id": candidate_id,
                "family": family,
                "params_json": json.dumps(params, ensure_ascii=False),
                "optimized_by": candidate.get("optimized_by", "manual_grid"),
                "feature_count": len(feature_cols),
                "cv_MAE_mean": fold_df["MAE"].mean(),
                "cv_RMSE_mean": fold_df["RMSE"].mean(),
                "cv_MAPE_mean": fold_df["MAPE_pct"].mean(),
                "cv_R2_mean": fold_df["R2"].mean(),
                "cv_MAE_std": fold_df["MAE"].std(),
                "cv_RMSE_std": fold_df["RMSE"].std(),
            }
        )

    results = pd.DataFrame(rows).sort_values(["cv_RMSE_mean", "cv_MAE_mean"]).reset_index(drop=True)
    return results, results.iloc[0].to_dict()


def cross_validate_sarimax(
    train_df: pd.DataFrame,
    exog_cols: list[str],
) -> tuple[pd.DataFrame, dict[str, object] | None]:
    if not HAS_STATSMODELS or not exog_cols:
        return pd.DataFrame(), None

    y_train = train_df["target_logret_t1"]
    exog_train = train_df[exog_cols]
    splits = 3
    tscv = TimeSeriesSplit(n_splits=splits)
    rows = []

    for candidate_id, candidate in enumerate(sarimax_candidates(), start=1):
        params = dict(candidate["params"])
        fold_metrics = []

        for fold_id, (idx_fit, idx_val) in enumerate(tscv.split(exog_train), start=1):
            try:
                model = SARIMAX(
                    y_train.iloc[idx_fit],
                    exog=exog_train.iloc[idx_fit],
                    order=params["order"],
                    trend=params.get("trend", "c"),
                    enforce_stationarity=False,
                    enforce_invertibility=False,
                )
                results = model.fit(disp=False)
                pred_val_logret = np.asarray(results.forecast(steps=len(idx_val), exog=exog_train.iloc[idx_val]), dtype=float)
                base_price_val = train_df.iloc[idx_val]["precio_oro"].to_numpy(dtype=float)
                pred_val_price = base_price_val * np.exp(pred_val_logret)
                y_val_price = train_df.iloc[idx_val]["target_precio_oro_t1"].to_numpy(dtype=float)
                metrics = regression_metrics(y_val_price, pred_val_price)
            except Exception:
                metrics = {
                    "MAE": float("inf"),
                    "RMSE": float("inf"),
                    "MAPE_pct": float("inf"),
                    "R2": float("-inf"),
                }
            metrics["fold"] = fold_id
            fold_metrics.append(metrics)

        fold_df = pd.DataFrame(fold_metrics)
        rows.append(
            {
                "candidate_id": candidate_id,
                "family": "SARIMAX",
                "params_json": json.dumps(params, ensure_ascii=False),
                "feature_count": len(exog_cols),
                "cv_MAE_mean": fold_df["MAE"].mean(),
                "cv_RMSE_mean": fold_df["RMSE"].mean(),
                "cv_MAPE_mean": fold_df["MAPE_pct"].mean(),
                "cv_R2_mean": fold_df["R2"].mean(),
                "cv_MAE_std": fold_df["MAE"].std(),
                "cv_RMSE_std": fold_df["RMSE"].std(),
            }
        )

    results = pd.DataFrame(rows).sort_values(["cv_RMSE_mean", "cv_MAE_mean"]).reset_index(drop=True)
    if results.empty:
        return results, None
    return results, results.iloc[0].to_dict()


def fit_model_by_spec(
    train_df: pd.DataFrame,
    feature_cols: list[str],
    sarimax_exog_cols: list[str],
    model_spec: dict[str, object],
):
    family = str(model_spec["family"])
    params = json.loads(model_spec["params_json"]) if isinstance(model_spec["params_json"], str) else model_spec["params_json"]

    if family == "SARIMAX":
        model = SARIMAX(
            train_df["target_logret_t1"],
            exog=train_df[sarimax_exog_cols],
            order=tuple(params["order"]),
            trend=params.get("trend", "c"),
            enforce_stationarity=False,
            enforce_invertibility=False,
        )
        return model.fit(disp=False)

    model = make_ml_model(family, params)
    model.fit(train_df[feature_cols], train_df["target_logret_t1"])
    return model


def predict_with_model(
    fitted_model,
    family: str,
    df_part: pd.DataFrame,
    feature_cols: list[str],
    sarimax_exog_cols: list[str],
) -> np.ndarray:
    pred_logret = predict_logret_with_model(fitted_model, family, df_part, feature_cols, sarimax_exog_cols)
    base_price = df_part["precio_oro"].to_numpy(dtype=float)
    return base_price * np.exp(pred_logret)


def predict_logret_with_model(
    fitted_model,
    family: str,
    df_part: pd.DataFrame,
    feature_cols: list[str],
    sarimax_exog_cols: list[str],
) -> np.ndarray:
    if family == "SARIMAX":
        return np.asarray(fitted_model.forecast(steps=len(df_part), exog=df_part[sarimax_exog_cols]), dtype=float)
    else:
        return np.asarray(fitted_model.predict(df_part[feature_cols]), dtype=float)


def evaluate_best_of_each_family(
    train_df: pd.DataFrame,
    test_df: pd.DataFrame,
    feature_cols: list[str],
    sarimax_exog_cols: list[str],
    cv_results: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, np.ndarray]]:
    rows = []
    predictions = {}

    for family in cv_results["family"].unique():
        family_best = cv_results[cv_results["family"] == family].iloc[0].to_dict()
        fitted = fit_model_by_spec(train_df, feature_cols, sarimax_exog_cols, family_best)
        pred_test_price = predict_with_model(fitted, family, test_df, feature_cols, sarimax_exog_cols)
        metrics = regression_metrics(test_df["target_precio_oro_t1"], pred_test_price)
        rows.append(
            {
                "family": family,
                "selection_metric_cv_rmse": family_best["cv_RMSE_mean"],
                "selection_metric_cv_mae": family_best["cv_MAE_mean"],
                "params_json": family_best["params_json"],
                **metrics,
            }
        )
        predictions[family] = pred_test_price

    results = pd.DataFrame(rows).sort_values(["selection_metric_cv_rmse", "RMSE", "MAE"]).reset_index(drop=True)
    return results, predictions


def build_raw_scenario_values(
    base_df: pd.DataFrame,
    future_dates: pd.DatetimeIndex,
    scenario_name: str,
) -> pd.DataFrame:
    raw_cols = [col for col in base_df.columns if col not in {"fecha", "precio_oro"}]
    scenario_df = pd.DataFrame({"fecha": future_dates})
    last_row = base_df.iloc[-1]

    annual_drifts = {
        "base": {
            "precio_plata": 0.040,
            "treasury_10y": -0.010,
            "breakeven_10y": 0.005,
            "cpi_us": 0.030,
            "dollar_index": -0.004,
            "fed_funds": -0.010,
            "vix": 0.003,
            "wti_oil": 0.015,
            "sp500": 0.025,
            "usd_cop": 0.012,
            "ni_precio_recursos": 0.015,
            "ni_precio_reservas": 0.015,
        },
        "crisis_alta_incertidumbre": {
            "precio_plata": 0.080,
            "treasury_10y": -0.050,
            "breakeven_10y": 0.035,
            "cpi_us": 0.045,
            "dollar_index": -0.020,
            "fed_funds": -0.040,
            "vix": 0.060,
            "wti_oil": 0.020,
            "sp500": -0.010,
            "usd_cop": 0.050,
            "ni_precio_recursos": 0.020,
            "ni_precio_reservas": 0.020,
        },
        "optimista": {
            "precio_plata": 0.025,
            "treasury_10y": 0.012,
            "breakeven_10y": -0.003,
            "cpi_us": 0.022,
            "dollar_index": 0.008,
            "fed_funds": 0.005,
            "vix": -0.040,
            "wti_oil": 0.010,
            "sp500": 0.045,
            "usd_cop": -0.010,
            "ni_precio_recursos": 0.010,
            "ni_precio_reservas": 0.010,
        },
    }
    drift_map = annual_drifts.get(scenario_name, annual_drifts["base"])

    for col in raw_cols:
        last_value = pd.to_numeric(last_row.get(col, np.nan), errors="coerce")
        if pd.isna(last_value):
            scenario_df[col] = np.nan
            continue
        if col == "tasa_real_aprox":
            continue

        drift = drift_map.get(col, 0.0)
        current = float(last_value)
        values = []
        for _ in future_dates:
            if col in {"precio_plata", "cpi_us", "dollar_index", "vix", "wti_oil", "sp500", "usd_cop"} or col.startswith("ni_"):
                current = current * (1 + drift / 252.0)
            else:
                current = current + drift / 252.0
            values.append(current)
        scenario_df[col] = values

    if "tasa_real_aprox" in raw_cols:
        if {"treasury_10y", "breakeven_10y"}.issubset(scenario_df.columns):
            scenario_df["tasa_real_aprox"] = scenario_df["treasury_10y"] - scenario_df["breakeven_10y"]
        else:
            scenario_df["tasa_real_aprox"] = float(last_row.get("tasa_real_aprox", np.nan))

    return scenario_df


def apply_scenario_return_rules(pred_logret: float, scenario_name: str) -> float:
    config = {
        "base": {"bias": 0.00005, "floor": -0.00005, "cap": 0.00180},
        "crisis_alta_incertidumbre": {"bias": 0.00018, "floor": 0.00002, "cap": 0.00250},
        "optimista": {"bias": 0.00000, "floor": -0.00012, "cap": 0.00120},
    }.get(scenario_name, {"bias": 0.0, "floor": -0.00050, "cap": 0.00180})

    adjusted = pred_logret + config["bias"]
    adjusted = max(adjusted, config["floor"])
    adjusted = min(adjusted, config["cap"])
    return adjusted


def estimate_monte_carlo_sigma(
    df_gold: pd.DataFrame,
    test_df: pd.DataFrame,
    pred_test_price: np.ndarray,
) -> float:
    realized_logret = np.log(df_gold["precio_oro"] / df_gold["precio_oro"].shift(1)).dropna()
    realized_sigma = float(realized_logret.tail(252).std()) if not realized_logret.empty else 0.006

    actual_test_logret = np.log(test_df["target_precio_oro_t1"].to_numpy(dtype=float) / test_df["precio_oro"].to_numpy(dtype=float))
    pred_test_logret = np.log(np.asarray(pred_test_price, dtype=float) / test_df["precio_oro"].to_numpy(dtype=float))
    residual_sigma = float(np.nanstd(actual_test_logret - pred_test_logret))

    sigma = max(realized_sigma, residual_sigma, 0.0045)
    return float(sigma)


def monte_carlo_config(scenario_name: str) -> dict[str, float]:
    return {
        "base": {"sigma_mult": 1.00, "jump_prob": 0.004, "jump_mean": 0.0000, "jump_std": 0.0045, "phi": 0.30},
        "crisis_alta_incertidumbre": {"sigma_mult": 1.45, "jump_prob": 0.010, "jump_mean": 0.0008, "jump_std": 0.0080, "phi": 0.35},
        "optimista": {"sigma_mult": 0.80, "jump_prob": 0.003, "jump_mean": -0.0001, "jump_std": 0.0035, "phi": 0.25},
    }.get(scenario_name, {"sigma_mult": 1.00, "jump_prob": 0.004, "jump_mean": 0.0000, "jump_std": 0.0045, "phi": 0.30})


def recursive_forecast_ml(
    model,
    base_df: pd.DataFrame,
    feature_cols: list[str],
    end_date: str | pd.Timestamp,
    scenario_name: str,
) -> pd.DataFrame:
    extended = base_df.copy().sort_values("fecha").reset_index(drop=True)
    future_dates = pd.date_range(extended["fecha"].max() + BDay(1), pd.Timestamp(end_date), freq="B")
    if len(future_dates) == 0:
        return pd.DataFrame(columns=["fecha", "precio_oro_predicho", "retorno_log_predicho", "escenario"])

    exogenous_cols = [col for col in extended.columns if col not in {"fecha", "precio_oro"}]
    scenario_raw = build_raw_scenario_values(extended, future_dates, scenario_name)
    forecasts = []

    for row_idx, future_date in enumerate(future_dates):
        feature_window = extended.tail(90).copy()
        featured_window = engineer_features(feature_window, add_targets=False)
        last_feature_row = featured_window.iloc[-1]
        X_next = (
            last_feature_row[feature_cols]
            .to_frame()
            .T.apply(pd.to_numeric, errors="coerce")
            .replace([np.inf, -np.inf], np.nan)
            .ffill(axis=1)
        )
        pred_logret = float(model.predict(X_next)[0])
        pred_logret = apply_scenario_return_rules(pred_logret, scenario_name)
        base_price = float(extended.iloc[-1]["precio_oro"])
        pred_price = float(base_price * np.exp(pred_logret))

        new_row = {"fecha": future_date, "precio_oro": pred_price}
        for col in exogenous_cols:
            new_row[col] = scenario_raw.iloc[row_idx][col] if col in scenario_raw.columns else extended[col].iloc[-1]

        extended = pd.concat([extended, pd.DataFrame([new_row])], ignore_index=True)
        forecasts.append(
            {
                "fecha": future_date,
                "precio_oro_predicho": pred_price,
                "retorno_log_predicho": pred_logret,
                "escenario": scenario_name,
            }
        )

    forecast_df = pd.DataFrame(forecasts)
    forecast_df["anio"] = forecast_df["fecha"].dt.year
    forecast_df["mes"] = forecast_df["fecha"].dt.month
    forecast_df["dia"] = forecast_df["fecha"].dt.day
    return forecast_df


def build_sarimax_future_exog(
    base_df: pd.DataFrame,
    end_date: str | pd.Timestamp,
    exog_cols: list[str],
    scenario_name: str,
) -> tuple[pd.DatetimeIndex, pd.DataFrame]:
    last_date = base_df["fecha"].max()
    future_target_dates = pd.date_range(last_date + BDay(1), pd.Timestamp(end_date), freq="B")
    if len(future_target_dates) == 0:
        return future_target_dates, pd.DataFrame(columns=exog_cols)

    scenario_raw_future = build_raw_scenario_values(base_df, future_target_dates, scenario_name)
    raw_cols = [col for col in base_df.columns if col not in {"fecha", "precio_oro"}]
    history_tail = base_df[["fecha"] + raw_cols].tail(30).copy()
    raw_history = pd.concat([history_tail, scenario_raw_future], ignore_index=True).drop_duplicates(subset=["fecha"], keep="last")
    raw_history = raw_history.sort_values("fecha").reset_index(drop=True)

    last_actual_row = base_df.iloc[-1]
    origin_dates = pd.DatetimeIndex([last_actual_row["fecha"]]).append(future_target_dates[:-1])
    raw_indexed = raw_history.set_index("fecha")
    origin_raw = raw_indexed.reindex(origin_dates).ffill()

    future_exog = pd.DataFrame({"fecha": origin_dates})
    for col in raw_cols:
        if col in origin_raw.columns:
            future_exog[col] = pd.to_numeric(origin_raw[col], errors="coerce")
            shifted_series = raw_indexed[col].reindex(origin_dates).shift(1)
            future_exog[f"{col}_lag_1"] = pd.to_numeric(shifted_series, errors="coerce")
            future_exog[f"{col}_chg_1"] = pd.to_numeric(raw_indexed[col].reindex(origin_dates).pct_change(1), errors="coerce")

    future_exog["anio"] = future_exog["fecha"].dt.year
    future_exog["mes"] = future_exog["fecha"].dt.month
    future_exog["trimestre"] = future_exog["fecha"].dt.quarter
    future_exog["dia_semana"] = future_exog["fecha"].dt.dayofweek
    future_exog["mes_inicio"] = future_exog["fecha"].dt.is_month_start.astype(int)
    future_exog["mes_fin"] = future_exog["fecha"].dt.is_month_end.astype(int)
    future_exog["mes_sin"] = np.sin(2 * np.pi * future_exog["mes"] / 12.0)
    future_exog["mes_cos"] = np.cos(2 * np.pi * future_exog["mes"] / 12.0)
    future_exog["dia_semana_sin"] = np.sin(2 * np.pi * future_exog["dia_semana"] / 7.0)
    future_exog["dia_semana_cos"] = np.cos(2 * np.pi * future_exog["dia_semana"] / 7.0)

    for col in exog_cols:
        if col not in future_exog.columns:
            future_exog[col] = np.nan

    future_exog = future_exog[["fecha"] + exog_cols].replace([np.inf, -np.inf], np.nan).ffill().bfill()
    return future_target_dates, future_exog


def forecast_with_sarimax(
    fitted_model,
    base_df: pd.DataFrame,
    end_date: str | pd.Timestamp,
    exog_cols: list[str],
    scenario_name: str,
) -> pd.DataFrame:
    future_target_dates, future_exog = build_sarimax_future_exog(base_df, end_date, exog_cols, scenario_name)
    if len(future_target_dates) == 0:
        return pd.DataFrame(columns=["fecha", "precio_oro_predicho", "retorno_log_predicho", "escenario"])

    pred_logrets = np.asarray(fitted_model.forecast(steps=len(future_target_dates), exog=future_exog[exog_cols]), dtype=float)

    prices = []
    current_price = float(base_df["precio_oro"].iloc[-1])
    adjusted_returns = []
    for pred_logret in pred_logrets:
        adj_ret = apply_scenario_return_rules(float(pred_logret), scenario_name)
        current_price = float(current_price * np.exp(adj_ret))
        adjusted_returns.append(adj_ret)
        prices.append(current_price)

    forecast_df = pd.DataFrame(
        {
            "fecha": future_target_dates,
            "precio_oro_predicho": prices,
            "retorno_log_predicho": adjusted_returns,
            "escenario": scenario_name,
        }
    )
    forecast_df["anio"] = forecast_df["fecha"].dt.year
    forecast_df["mes"] = forecast_df["fecha"].dt.month
    forecast_df["dia"] = forecast_df["fecha"].dt.day
    return forecast_df


def run_monte_carlo_from_base_forecast(
    base_forecast_df: pd.DataFrame,
    scenario_name: str,
    shock_sigma: float,
    n_sims: int = MONTE_CARLO_SIMULATIONS,
    sample_paths: int = MONTE_CARLO_SAMPLE_PATHS,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if base_forecast_df.empty:
        empty_summary = pd.DataFrame(
            columns=[
                "fecha",
                "escenario",
                "precio_oro_p10",
                "precio_oro_p50",
                "precio_oro_p90",
                "precio_oro_media",
                "retorno_log_base",
                "anio",
                "mes",
                "dia",
            ]
        )
        empty_paths = pd.DataFrame(columns=["fecha", "escenario", "simulacion", "precio_oro_simulado"])
        return empty_summary, empty_paths

    cfg = monte_carlo_config(scenario_name)
    rng = np.random.default_rng(RANDOM_STATE + abs(hash(scenario_name)) % 10000)
    base_returns = base_forecast_df["retorno_log_predicho"].to_numpy(dtype=float)
    dates = base_forecast_df["fecha"].to_numpy()
    start_price = float(base_forecast_df["precio_oro_predicho"].iloc[0] / np.exp(base_returns[0]))

    n_steps = len(base_returns)
    simulated_prices = np.zeros((n_sims, n_steps), dtype=float)
    simulated_returns = np.zeros((n_sims, n_steps), dtype=float)

    for sim in range(n_sims):
        current_price = start_price
        prev_shock = 0.0
        for t in range(n_steps):
            innovation = rng.normal()
            correlated_shock = cfg["phi"] * prev_shock + math.sqrt(max(1.0 - cfg["phi"] ** 2, 1e-8)) * innovation
            jump = 0.0
            if rng.random() < cfg["jump_prob"]:
                jump = rng.normal(cfg["jump_mean"], cfg["jump_std"])
            stochastic_ret = base_returns[t] + correlated_shock * shock_sigma * cfg["sigma_mult"] + jump
            stochastic_ret = float(np.clip(stochastic_ret, -0.05, 0.05))
            current_price = float(current_price * np.exp(stochastic_ret))
            simulated_returns[sim, t] = stochastic_ret
            simulated_prices[sim, t] = current_price
            prev_shock = correlated_shock

    summary_df = pd.DataFrame(
        {
            "fecha": pd.to_datetime(dates),
            "escenario": scenario_name,
            "precio_oro_p10": np.quantile(simulated_prices, 0.10, axis=0),
            "precio_oro_p50": np.quantile(simulated_prices, 0.50, axis=0),
            "precio_oro_p90": np.quantile(simulated_prices, 0.90, axis=0),
            "precio_oro_media": simulated_prices.mean(axis=0),
            "retorno_log_base": base_returns,
        }
    )
    summary_df["precio_oro_predicho"] = summary_df["precio_oro_p50"]
    summary_df["anio"] = summary_df["fecha"].dt.year
    summary_df["mes"] = summary_df["fecha"].dt.month
    summary_df["dia"] = summary_df["fecha"].dt.day

    sample_keep = min(sample_paths, n_sims)
    sample_ids = np.linspace(0, n_sims - 1, sample_keep, dtype=int)
    path_rows = []
    for sim_id in sample_ids:
        path_rows.append(
            pd.DataFrame(
                {
                    "fecha": pd.to_datetime(dates),
                    "escenario": scenario_name,
                    "simulacion": int(sim_id) + 1,
                    "precio_oro_simulado": simulated_prices[sim_id],
                    "retorno_log_simulado": simulated_returns[sim_id],
                }
            )
        )
    sample_paths_df = pd.concat(path_rows, ignore_index=True)
    return summary_df, sample_paths_df


def build_monthly_forecast_summary(forecast_2026_df: pd.DataFrame) -> pd.DataFrame:
    month_names = {5: "mayo", 6: "junio", 7: "julio"}
    return (
        forecast_2026_df.assign(nombre_mes=forecast_2026_df["mes"].map(month_names))
        .groupby(["escenario", "anio", "mes", "nombre_mes"], as_index=False)
        .agg(
            precio_promedio_usd_oz=("precio_oro_predicho", "mean"),
            precio_minimo_usd_oz=("precio_oro_predicho", "min"),
            precio_maximo_usd_oz=("precio_oro_predicho", "max"),
            observaciones=("precio_oro_predicho", "size"),
        )
        .sort_values(["escenario", "anio", "mes"])
        .reset_index(drop=True)
    )


def plot_predictions(pred_df: pd.DataFrame, best_family: str) -> Path:
    fig, axes = plt.subplots(2, 1, figsize=(14, 9), sharex=False)

    axes[0].plot(pred_df["fecha_objetivo"], pred_df["real_t1"], label="Real", color="black", linewidth=1.5)
    axes[0].plot(pred_df["fecha_objetivo"], pred_df["predicho_t1"], label="Predicho", color="crimson", linewidth=1.5)
    axes[0].set_title(f"Prediccion vs valor real en test - modelo ganador: {best_family}")
    axes[0].set_xlabel("Fecha objetivo")
    axes[0].set_ylabel("Precio del oro (USD/oz)")
    axes[0].grid(alpha=0.25)
    axes[0].legend()

    axes[1].scatter(pred_df["real_t1"], pred_df["predicho_t1"], color="steelblue", alpha=0.65, s=18)
    min_val = min(float(pred_df["real_t1"].min()), float(pred_df["predicho_t1"].min()))
    max_val = max(float(pred_df["real_t1"].max()), float(pred_df["predicho_t1"].max()))
    axes[1].plot([min_val, max_val], [min_val, max_val], color="darkred", linestyle="--")
    axes[1].set_title("Dispersion entre valor real y predicho")
    axes[1].set_xlabel("Valor real")
    axes[1].set_ylabel("Valor predicho")
    axes[1].grid(alpha=0.25)

    fig.tight_layout()
    fig.savefig(PLOT_PRED, dpi=300)
    plt.close(fig)
    return PLOT_PRED


def plot_test_errors(pred_df: pd.DataFrame) -> Path:
    fig, axes = plt.subplots(2, 1, figsize=(14, 8), sharex=False)
    axes[0].plot(pred_df["fecha_objetivo"], pred_df["error"], color="firebrick", linewidth=1.1)
    axes[0].axhline(0, color="black", linewidth=0.9, linestyle="--")
    axes[0].set_title("Error de prediccion en test: real menos predicho")
    axes[0].set_xlabel("Fecha objetivo")
    axes[0].set_ylabel("Error (USD/oz)")
    axes[0].grid(alpha=0.25)

    sns.histplot(pred_df["error"], bins=35, kde=True, ax=axes[1], color="slateblue")
    axes[1].axvline(0, color="black", linewidth=0.9, linestyle="--")
    axes[1].set_title("Distribucion de errores en test")
    axes[1].set_xlabel("Error (USD/oz)")
    axes[1].set_ylabel("Frecuencia")
    axes[1].grid(alpha=0.25)

    fig.tight_layout()
    fig.savefig(PLOT_ERRORS, dpi=300)
    plt.close(fig)
    return PLOT_ERRORS


def plot_forecast_may_jul_2026(df_gold: pd.DataFrame, forecast_df: pd.DataFrame) -> Path:
    history_start = max(df_gold["fecha"].min(), SHORT_FORECAST_START - pd.DateOffset(years=2))
    history_subset = df_gold[df_gold["fecha"] >= history_start].copy()

    fig, ax = plt.subplots(figsize=(14, 6))
    ax.plot(history_subset["fecha"], history_subset["precio_oro"], color="black", linewidth=1.4, label="Historico")
    for scenario in SCENARIO_ORDER:
        subset = forecast_df[forecast_df["escenario"] == scenario].copy()
        if subset.empty:
            continue
        ax.fill_between(
            subset["fecha"],
            subset["precio_oro_p10"],
            subset["precio_oro_p90"],
            color=SCENARIO_COLORS.get(scenario, "gray"),
            alpha=0.15,
        )
        ax.plot(
            subset["fecha"],
            subset["precio_oro_p50"],
            color=SCENARIO_COLORS.get(scenario, "gray"),
            linewidth=1.8,
            label=f"{scenario} (mediana)",
        )
    ax.axvspan(SHORT_FORECAST_START, SHORT_FORECAST_END, color="lightsteelblue", alpha=0.20)
    ax.set_title("Historico del oro y pronostico Monte Carlo para mayo, junio y julio de 2026")
    ax.set_xlabel("Fecha")
    ax.set_ylabel("Precio del oro (USD/oz)")
    ax.grid(alpha=0.25)
    ax.legend()
    fig.tight_layout()
    fig.savefig(PLOT_FORECAST_2026, dpi=300)
    plt.close(fig)
    return PLOT_FORECAST_2026


def plot_forecast_to_2035(
    df_gold: pd.DataFrame,
    forecast_df: pd.DataFrame,
) -> Path:
    recent_history = df_gold[df_gold["fecha"] >= (df_gold["fecha"].max() - pd.DateOffset(years=3))].copy()
    fig, ax = plt.subplots(figsize=(14, 6))

    ax.plot(recent_history["fecha"], recent_history["precio_oro"], color="black", linewidth=1.3, label="Historico reciente")
    for scenario in SCENARIO_ORDER:
        subset = forecast_df[forecast_df["escenario"] == scenario].copy()
        if subset.empty:
            continue
        color = SCENARIO_COLORS.get(scenario, "gray")
        ax.fill_between(subset["fecha"], subset["precio_oro_p10"], subset["precio_oro_p90"], color=color, alpha=0.15)
        ax.plot(subset["fecha"], subset["precio_oro_p50"], color=color, linewidth=1.8, label=f"{scenario} (mediana)")
    ax.set_title("Historico reciente y escenarios Monte Carlo del oro hasta 2035")
    ax.set_xlabel("Fecha")
    ax.set_ylabel("Precio del oro (USD/oz)")
    ax.grid(alpha=0.25)
    ax.legend()

    fig.tight_layout()
    fig.savefig(PLOT_FORECAST_2035, dpi=300)
    plt.close(fig)
    return PLOT_FORECAST_2035


def plot_monte_carlo_sample_paths(
    forecast_df: pd.DataFrame,
    sample_paths_df: pd.DataFrame,
) -> Path:
    fig, ax = plt.subplots(figsize=(14, 6))

    for scenario in SCENARIO_ORDER:
        color = SCENARIO_COLORS.get(scenario, "gray")
        subset_paths = sample_paths_df[sample_paths_df["escenario"] == scenario].copy()
        if not subset_paths.empty:
            for _, sim_df in subset_paths.groupby("simulacion"):
                ax.plot(sim_df["fecha"], sim_df["precio_oro_simulado"], color=color, linewidth=0.8, alpha=0.18)

        subset_summary = forecast_df[forecast_df["escenario"] == scenario].copy()
        if not subset_summary.empty:
            ax.plot(
                subset_summary["fecha"],
                subset_summary["precio_oro_p50"],
                color=color,
                linewidth=2.0,
                label=f"{scenario} (mediana)",
            )

    ax.set_title("Trayectorias simuladas de muestra y mediana por escenario hasta diciembre de 2035")
    ax.set_xlabel("Fecha")
    ax.set_ylabel("Precio del oro pronosticado (USD/oz)")
    ax.grid(alpha=0.25)
    ax.legend()

    fig.tight_layout()
    fig.savefig(PLOT_FORECAST_2035_PATHS, dpi=300)
    plt.close(fig)
    return PLOT_FORECAST_2035_PATHS


def build_single_trajectory(
    sample_paths_df: pd.DataFrame,
    forecast_2035_df: pd.DataFrame,
    scenario_name: str = "base",
    simulation_id: int = 1,
) -> pd.DataFrame:
    trajectory = sample_paths_df[
        (sample_paths_df["escenario"] == scenario_name)
        & (sample_paths_df["simulacion"] == simulation_id)
    ].copy()
    median_path = forecast_2035_df[forecast_2035_df["escenario"] == scenario_name][
        ["fecha", "precio_oro_p50", "precio_oro_p10", "precio_oro_p90", "retorno_log_base"]
    ].copy()

    single_df = trajectory.merge(median_path, on="fecha", how="left")
    single_df = single_df.rename(
        columns={
            "precio_oro_p50": "mediana_escenario_base",
            "precio_oro_p10": "banda_p10_base",
            "precio_oro_p90": "banda_p90_base",
        }
    )
    single_df["choque_logret"] = single_df["retorno_log_simulado"] - single_df["retorno_log_base"]
    single_df["precio_inicial_estimado"] = single_df["precio_oro_simulado"].iloc[0] / np.exp(
        single_df["retorno_log_simulado"].iloc[0]
    )
    return single_df


def plot_single_monte_carlo_trajectory(df_gold: pd.DataFrame, single_df: pd.DataFrame) -> Path:
    recent_history = df_gold[df_gold["fecha"] >= (df_gold["fecha"].max() - pd.DateOffset(years=2))].copy()
    fig, ax = plt.subplots(figsize=(14, 6))

    ax.plot(recent_history["fecha"], recent_history["precio_oro"], color="black", linewidth=1.3, label="Historico reciente")
    ax.fill_between(
        single_df["fecha"],
        single_df["banda_p10_base"],
        single_df["banda_p90_base"],
        color="darkorange",
        alpha=0.13,
        label="Banda P10-P90 del escenario base",
    )
    ax.plot(
        single_df["fecha"],
        single_df["mediana_escenario_base"],
        color="darkorange",
        linewidth=1.8,
        label="Mediana Monte Carlo base",
    )
    ax.plot(
        single_df["fecha"],
        single_df["precio_oro_simulado"],
        color="crimson",
        linewidth=1.2,
        label="Una trayectoria simulada",
    )
    ax.set_title("Una trayectoria Monte Carlo del precio del oro - escenario base")
    ax.set_xlabel("Fecha")
    ax.set_ylabel("Precio del oro (USD/oz)")
    ax.grid(alpha=0.25)
    ax.legend()
    fig.tight_layout()
    fig.savefig(PLOT_SINGLE_TRAJECTORY, dpi=300)
    plt.close(fig)
    return PLOT_SINGLE_TRAJECTORY


def build_importance_from_model(
    fitted_model,
    family: str,
    feature_cols: list[str],
    sarimax_exog_cols: list[str],
) -> pd.DataFrame:
    if family == "SARIMAX":
        rows = []
        for col in sarimax_exog_cols:
            if col in fitted_model.params.index:
                rows.append({"feature": col, "importance": abs(float(fitted_model.params[col]))})
        return pd.DataFrame(rows).sort_values("importance", ascending=False).reset_index(drop=True)

    return pd.DataFrame(
        {
            "feature": feature_cols,
            "importance": getattr(fitted_model, "feature_importances_", np.zeros(len(feature_cols))),
        }
    ).sort_values("importance", ascending=False).reset_index(drop=True)


def plot_feature_importance(importance_df: pd.DataFrame, best_family: str) -> tuple[Path, pd.DataFrame]:
    top_importance = importance_df.head(20).sort_values("importance", ascending=True)
    fig, ax = plt.subplots(figsize=(11, 8))
    ax.barh(top_importance["feature"], top_importance["importance"], color="teal", alpha=0.9)
    ax.set_title(f"Importancia de variables - modelo ganador: {best_family}")
    ax.set_xlabel("Importancia")
    ax.set_ylabel("Variable")
    ax.grid(alpha=0.2, axis="x")
    fig.tight_layout()
    fig.savefig(PLOT_IMPORTANCE, dpi=300)
    plt.close(fig)
    return PLOT_IMPORTANCE, importance_df


def compute_shap_outputs(
    fitted_model,
    family: str,
    X_train: pd.DataFrame,
) -> tuple[Path | None, pd.DataFrame | None]:
    if not HAS_SHAP or family == "SARIMAX":
        return None, None

    sample_size = min(400, len(X_train))
    X_sample = X_train.tail(sample_size).copy()
    try:
        explainer = shap.TreeExplainer(fitted_model)
        shap_values = explainer.shap_values(X_sample)
        shap_df = pd.DataFrame(
            {
                "feature": X_sample.columns,
                "mean_abs_shap": np.abs(shap_values).mean(axis=0),
            }
        ).sort_values("mean_abs_shap", ascending=False)

        plt.figure(figsize=(11, 7))
        shap.summary_plot(shap_values, X_sample, show=False, max_display=15)
        plt.tight_layout()
        plt.savefig(PLOT_SHAP, dpi=300, bbox_inches="tight")
        plt.close()
        return PLOT_SHAP, shap_df
    except Exception as exc:
        print(f"\nSHAP no pudo generarse automaticamente: {exc}")
        return None, None


def technical_interpretation(
    metadata: dict[str, str],
    used_exogenous: list[str],
    cv_results: pd.DataFrame,
    test_results: pd.DataFrame,
    best_family: str,
    best_params_json: str,
    importance_df: pd.DataFrame,
    shap_df: pd.DataFrame | None,
    macro_warnings: list[str],
) -> str:
    best_cv = cv_results.iloc[0]
    best_test = test_results[test_results["family"] == best_family].iloc[0]
    top_features = importance_df.head(10)["feature"].tolist()
    shap_features = [] if shap_df is None else shap_df.head(10)["feature"].tolist()

    lines = [
        "REPORTE TECNICO - MODELO MULTIVARIADO PARA PRECIO DEL ORO",
        "",
        f"Archivo analizado: {metadata['archivo']}",
        f"Hoja de oro detectada automaticamente: {metadata['hoja_oro']}",
        f"Columna objetivo detectada: {metadata['columna_objetivo']}",
        "Objetivo supervisado: prediccion del log-retorno del siguiente precio de cierre del oro (t+1).",
        "Esto permite una modelacion mas estable y evita fuga de informacion al usar solo informacion disponible en t.",
        "",
        "Variables utilizadas:",
        "- Rezagos del precio del oro",
        "- Retornos, diferencias, medias moviles y volatilidad movil",
        "- Variables temporales",
        f"- Variables exogenas efectivamente utilizadas: {', '.join(used_exogenous) if used_exogenous else 'ninguna usable'}",
        "",
        "Comparacion de modelos realizada con validacion temporal:",
        "- XGBoost",
        "- RandomForest",
        "- ExtraTrees",
        f"- LightGBM: {'disponible' if HAS_LIGHTGBM else 'no disponible'}",
        f"- CatBoost: {'disponible' if HAS_CATBOOST else 'no disponible'}",
        "- HistGradientBoosting",
        f"- Optuna para busqueda automatica: {'disponible' if HAS_OPTUNA else 'no disponible'}",
        f"- SARIMAX con variables exogenas: {'evaluado' if ENABLE_SARIMAX_REFERENCE and HAS_STATSMODELS else 'no evaluado en la version rapida'}",
        "",
        f"Modelo seleccionado por mejor RMSE promedio en validacion temporal: {best_family}",
        f"RMSE promedio CV del mejor modelo: {best_cv['cv_RMSE_mean']:.4f}",
        f"MAE promedio CV del mejor modelo: {best_cv['cv_MAE_mean']:.4f}",
        f"R2 promedio CV del mejor modelo: {best_cv['cv_R2_mean']:.4f}",
        f"Hiperparametros / especificacion ganadora: {best_params_json}",
        "",
        "Metricas en test del modelo ganador:",
        f"- MAE: {best_test['MAE']:.4f}",
        f"- RMSE: {best_test['RMSE']:.4f}",
        f"- MAPE: {best_test['MAPE_pct']:.4f}%",
        f"- R2: {best_test['R2']:.4f}",
        "",
        "Lectura economica de los escenarios:",
        "- Escenario base: tasas reales algo mas bajas, dolar ligeramente mas debil y volatilidad contenida; el oro tiende a una senda de crecimiento moderado.",
        "- Escenario de crisis / alta incertidumbre: VIX mas alto, tasas reales mas bajas, dolar mas debil y mayor demanda de refugio; el oro tiende a apreciarse con mayor fuerza.",
        "- Escenario optimista: mejor apetito por riesgo, S&P 500 mas fuerte, VIX menor y dolar mas firme; el oro crece de forma mas lenta o se estabiliza.",
        "- Para el horizonte largo se usa simulacion Monte Carlo alrededor del retorno base del modelo, incorporando volatilidad y saltos por escenario.",
        "",
        f"Variables mas importantes segun el modelo ganador: {', '.join(top_features[:10])}",
    ]

    if shap_features:
        lines.append(f"Variables mas influyentes segun SHAP: {', '.join(shap_features[:10])}")
    else:
        lines.append("SHAP no se genero o no aplica al modelo ganador.")

    if macro_warnings:
        lines.extend(["", "Advertencias de descarga externa:"] + [f"- {item}" for item in macro_warnings])

    lines.extend(
        [
            "",
            "Conclusion tecnica:",
            "El flujo es reproducible, respeta el orden temporal, no usa train_test_split aleatorio y evita mezclar pasado y futuro.",
            "El horizonte corto se interpreta como pronostico condicional al ultimo bloque historico observado.",
            "El horizonte hasta 2035 debe leerse como simulacion por escenarios y no como un unico valor exacto determinista.",
        ]
    )

    return "\n".join(lines)


def format_excel_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    scenario_fills = {
        "base": PatternFill("solid", fgColor="FFF2CC"),
        "crisis_alta_incertidumbre": PatternFill("solid", fgColor="E2F0D9"),
        "optimista": PatternFill("solid", fgColor="D9EAF7"),
    }

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet.iter_rows(min_row=2):
            scenario_value = row[0].value if worksheet.max_column >= 1 else None
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value is not None:
                    cell.number_format = '#,##0.00'
                if hasattr(cell.value, "year") and hasattr(cell.value, "month") and hasattr(cell.value, "day"):
                    cell.number_format = "yyyy-mm-dd"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(scenario_value, str) and scenario_value in scenario_fills:
                for row_cell in row:
                    row_cell.fill = scenario_fills[scenario_value]

        for column_cells in worksheet.columns:
            max_length = 0
            col_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 28)

    workbook.save(path)


def build_scenario_summary_2035(forecast_2035_df: pd.DataFrame) -> pd.DataFrame:
    return (
        forecast_2035_df.groupby("escenario", as_index=False)
        .agg(
            fecha_inicio=("fecha", "min"),
            fecha_fin=("fecha", "max"),
            precio_inicio=("precio_oro_p50", "first"),
            precio_fin=("precio_oro_p50", "last"),
            precio_minimo=("precio_oro_p10", "min"),
            precio_maximo=("precio_oro_p90", "max"),
            crecimiento_pct=("precio_oro_p50", lambda s: ((s.iloc[-1] / s.iloc[0]) - 1.0) * 100.0),
        )
        .sort_values("escenario")
        .reset_index(drop=True)
    )


def save_outputs_to_excel(
    metadata: dict[str, str],
    df_gold: pd.DataFrame,
    df_macro: pd.DataFrame,
    df_model: pd.DataFrame,
    train_df: pd.DataFrame,
    test_df: pd.DataFrame,
    cv_results: pd.DataFrame,
    test_results: pd.DataFrame,
    pred_df: pd.DataFrame,
    forecast_2026_df: pd.DataFrame,
    forecast_2035_df: pd.DataFrame,
    monthly_summary_df: pd.DataFrame,
    importance_df: pd.DataFrame,
    shap_df: pd.DataFrame | None,
    scenario_summary_df: pd.DataFrame,
    sample_paths_df: pd.DataFrame,
    single_trajectory_df: pd.DataFrame,
) -> Path:
    metadata_df = pd.DataFrame({"campo": list(metadata.keys()), "valor": list(metadata.values())})
    eda_df = build_eda_table(df_gold).reset_index().rename(columns={"variable": "estadistico"})
    split_df = pd.DataFrame(
        {
            "conjunto": ["train", "test"],
            "observaciones": [len(train_df), len(test_df)],
            "fecha_min": [train_df["fecha"].min(), test_df["fecha"].min()],
            "fecha_max": [train_df["fecha"].max(), test_df["fecha"].max()],
        }
    )

    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
        metadata_df.to_excel(writer, sheet_name="fuentes_detectadas", index=False)
        eda_df.to_excel(writer, sheet_name="eda_oro", index=False)
        df_macro.to_excel(writer, sheet_name="macro_fred", index=False)
        split_df.to_excel(writer, sheet_name="particion_temporal", index=False)
        cv_results.to_excel(writer, sheet_name="comparacion_modelos_cv", index=False)
        test_results.to_excel(writer, sheet_name="metricas_test_modelos", index=False)
        pred_df.to_excel(writer, sheet_name="predicciones_test", index=False)
        monthly_summary_df.to_excel(writer, sheet_name="resumen_mensual_2026", index=False)
        scenario_summary_df.to_excel(writer, sheet_name="resumen_escenarios_2035", index=False)
        forecast_2026_df.to_excel(writer, sheet_name="pronostico_may_jul_2026", index=False)
        forecast_2035_df.to_excel(writer, sheet_name="pronostico_hasta_2035", index=False)
        sample_paths_df.to_excel(writer, sheet_name="mc_muestra_2035", index=False)
        single_trajectory_df.to_excel(writer, sheet_name="trayectoria_unica_mc", index=False)
        importance_df.to_excel(writer, sheet_name="importancia_variables", index=False)
        df_model.to_excel(writer, sheet_name="dataset_modelado", index=False)
        if shap_df is not None:
            shap_df.to_excel(writer, sheet_name="shap_importancia", index=False)

    format_excel_workbook(OUTPUT_EXCEL)
    return OUTPUT_EXCEL


def main() -> None:
    file_path = first_existing_file()
    df_gold, gold_metadata = load_gold_data(file_path)
    df_silver, silver_metadata = load_silver_data(file_path)
    df_ni, ni_metadata = load_ni_data(file_path)
    df_macro, macro_metadata, macro_warnings = fetch_fred_macro_data()

    metadata = {**gold_metadata, **silver_metadata, **ni_metadata, **macro_metadata}
    print_basic_eda(df_gold, metadata)
    create_basic_eda_plot(df_gold)

    df_model, base_forecast_df, used_exogenous = build_model_dataset(df_gold, df_silver, df_ni, df_macro)
    train_df, test_df = temporal_train_test_split(df_model)
    feature_cols = select_ml_feature_columns(df_model)
    sarimax_exog_cols = select_sarimax_exog_columns(df_model)

    print("\n=== MODELADO MULTIMODELO ===")
    print(f"Observaciones para modelado: {len(df_model)}")
    print(f"Train: {len(train_df)} | Test: {len(test_df)}")
    print(f"Numero de variables ML: {len(feature_cols)}")
    print(f"Numero de variables SARIMAX exogenas: {len(sarimax_exog_cols)}")
    print(f"Variables exogenas disponibles: {', '.join(used_exogenous) if used_exogenous else 'ninguna'}")

    ml_cv_results, _ = cross_validate_ml_models(train_df, feature_cols)
    sarimax_cv_results, _ = cross_validate_sarimax(train_df, sarimax_exog_cols) if ENABLE_SARIMAX_REFERENCE else (pd.DataFrame(), None)
    cv_results = pd.concat([ml_cv_results, sarimax_cv_results], ignore_index=True)
    cv_results = cv_results.sort_values(["cv_RMSE_mean", "cv_MAE_mean"]).reset_index(drop=True)
    best_spec = cv_results.iloc[0].to_dict()
    best_family = str(best_spec["family"])

    print("\n=== COMPARACION DE MODELOS EN VALIDACION TEMPORAL ===")
    print(cv_results.to_string(index=False))
    print(f"\nModelo ganador por CV: {best_family}")
    print(f"Especificacion ganadora: {best_spec['params_json']}")

    test_results, predictions_by_family = evaluate_best_of_each_family(train_df, test_df, feature_cols, sarimax_exog_cols, cv_results)
    print("\n=== RESULTADOS EN TEST POR FAMILIA ===")
    print(test_results.to_string(index=False))

    pred_test_price = predictions_by_family[best_family]
    best_test_metrics = test_results[test_results["family"] == best_family].iloc[0].to_dict()

    pred_df = test_df[["fecha"]].copy()
    pred_df["fecha_objetivo"] = pred_df["fecha"] + BDay(1)
    pred_df["precio_actual_t"] = test_df["precio_oro"].values
    pred_df["real_t1"] = test_df["target_precio_oro_t1"].values
    pred_df["predicho_t1"] = pred_test_price
    pred_df["error"] = pred_df["real_t1"] - pred_df["predicho_t1"]
    pred_df["error_abs"] = np.abs(pred_df["error"])
    pred_df.to_csv(PREDICTIONS_CSV, index=False, encoding="utf-8-sig")

    mc_sigma = estimate_monte_carlo_sigma(df_gold, test_df, pred_test_price)
    full_model = fit_model_by_spec(df_model, feature_cols, sarimax_exog_cols, best_spec)
    try:
        joblib.dump(
            {
                "model": full_model,
                "best_family": best_family,
                "feature_cols": feature_cols,
                "sarimax_exog_cols": sarimax_exog_cols,
                "best_spec": best_spec,
            },
            MODEL_FILE,
        )
    except Exception as exc:
        print(f"No se pudo guardar el modelo con joblib: {exc}")

    scenario_forecasts = []
    monte_carlo_samples = []
    for scenario_name in SCENARIO_ORDER:
        if best_family == "SARIMAX":
            deterministic_df = forecast_with_sarimax(full_model, base_forecast_df, LONG_FORECAST_END, sarimax_exog_cols, scenario_name)
        else:
            deterministic_df = recursive_forecast_ml(full_model, base_forecast_df, feature_cols, LONG_FORECAST_END, scenario_name)
        scenario_df, sample_paths_df = run_monte_carlo_from_base_forecast(
            deterministic_df,
            scenario_name,
            shock_sigma=mc_sigma,
            n_sims=MONTE_CARLO_SIMULATIONS,
            sample_paths=MONTE_CARLO_SAMPLE_PATHS,
        )
        scenario_forecasts.append(scenario_df)
        monte_carlo_samples.append(sample_paths_df)

    forecast_2035_df = pd.concat(scenario_forecasts, ignore_index=True)
    sample_paths_2035_df = pd.concat(monte_carlo_samples, ignore_index=True)
    forecast_2026_df = forecast_2035_df[
        (forecast_2035_df["fecha"] >= SHORT_FORECAST_START)
        & (forecast_2035_df["fecha"] <= SHORT_FORECAST_END)
    ].copy()
    forecast_2026_df.to_csv(FORECAST_2026_CSV, index=False, encoding="utf-8-sig")
    forecast_2035_df.to_csv(FORECAST_2035_CSV, index=False, encoding="utf-8-sig")

    single_trajectory_df = build_single_trajectory(sample_paths_2035_df, forecast_2035_df)
    single_trajectory_df.to_csv(SINGLE_TRAJECTORY_CSV, index=False, encoding="utf-8-sig")

    plot_predictions(pred_df, best_family)
    plot_test_errors(pred_df)
    plot_forecast_may_jul_2026(df_gold, forecast_2026_df)
    plot_forecast_to_2035(df_gold, forecast_2035_df)
    plot_monte_carlo_sample_paths(forecast_2035_df, sample_paths_2035_df)
    plot_single_monte_carlo_trajectory(df_gold, single_trajectory_df)

    importance_df = build_importance_from_model(full_model, best_family, feature_cols, sarimax_exog_cols)
    _, importance_df = plot_feature_importance(importance_df, best_family)
    shap_plot_path, shap_df = compute_shap_outputs(full_model, best_family, df_model[feature_cols])

    report_text = technical_interpretation(
        metadata,
        used_exogenous,
        cv_results,
        test_results,
        best_family,
        str(best_spec["params_json"]),
        importance_df,
        shap_df,
        macro_warnings,
    )
    REPORT_TXT.write_text(report_text, encoding="utf-8")

    monthly_summary_df = build_monthly_forecast_summary(forecast_2026_df)
    scenario_summary_df = build_scenario_summary_2035(forecast_2035_df)
    save_outputs_to_excel(
        metadata,
        df_gold,
        df_macro,
        df_model,
        train_df,
        test_df,
        cv_results,
        test_results,
        pred_df,
        forecast_2026_df,
        forecast_2035_df,
        monthly_summary_df,
        importance_df,
        shap_df,
        scenario_summary_df,
        sample_paths_2035_df,
        single_trajectory_df,
    )

    print("\n=== METRICAS DEL MODELO GANADOR EN TEST ===")
    for metric_name in ["MAE", "RMSE", "MAPE_pct", "R2"]:
        metric_value = float(best_test_metrics[metric_name])
        if metric_name == "MAPE_pct":
            print(f"{metric_name}: {metric_value:.4f}%")
        else:
            print(f"{metric_name}: {metric_value:.4f}")
    print(f"Sigma usada en Monte Carlo: {mc_sigma:.5f}")

    print("\n=== TOP 15 VARIABLES ===")
    print(importance_df.head(15).to_string(index=False))

    print("\n=== PRONOSTICO MAYO-JUNIO-JULIO 2026 ===")
    print(forecast_2026_df.head(20).to_string(index=False))

    print("\n=== ARCHIVOS GENERADOS ===")
    generated_files = [
        PLOT_SERIES,
        PLOT_PRED,
        PLOT_ERRORS,
        PLOT_FORECAST_2026,
        PLOT_FORECAST_2035,
        PLOT_FORECAST_2035_PATHS,
        PLOT_SINGLE_TRAJECTORY,
        PLOT_IMPORTANCE,
        PLOT_SHAP if shap_plot_path is not None else None,
        PREDICTIONS_CSV,
        FORECAST_2026_CSV,
        FORECAST_2035_CSV,
        SINGLE_TRAJECTORY_CSV,
        OUTPUT_EXCEL,
        MODEL_FILE,
        REPORT_TXT,
    ]
    for output in generated_files:
        if output is not None:
            print(output)


if __name__ == "__main__":
    main()
